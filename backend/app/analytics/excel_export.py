import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(ws, row=1, ncols=1):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def render_analytics_excel(
    summary: dict,
    stock_movements: list,
    top_parts: list,
) -> bytes:
    """summary: dict of AnalyticsSummary fields.
    stock_movements: list of dicts (from StockMovementOut).
    top_parts: list of dicts (from TopPartRow), assumed already sorted desc."""
    wb = Workbook()

    # ── Summary sheet ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Metric", "Value"])
    _style_header(ws, ncols=2)
    rows = [
        ("Total parts tracked", summary.get("total_parts", 0)),
        ("Low stock parts", summary.get("low_stock_parts", 0)),
        ("Out of stock parts", summary.get("out_of_stock_parts", 0)),
        ("Total stock value (KES)", float(summary.get("total_stock_value", 0))),
        ("Goods received qty (period)", summary.get("goods_received_qty", 0)),
        ("Goods received value (KES)", float(summary.get("goods_received_value", 0))),
        ("Manual stock added qty (period)", summary.get("manual_stock_added_qty", 0)),
        ("Manual stock added value (KES)", float(summary.get("manual_stock_added_value", 0))),
        ("Sales qty (period)", summary.get("sales_qty", 0)),
        ("Sales value (KES)", float(summary.get("sales_value", 0))),
        ("Total purchases value (KES)", float(summary.get("total_purchases_value", 0))),
        ("Total expenses (KES)", float(summary.get("total_expenses", 0))),
        ("Net stock movement value (KES)", float(summary.get("net_movement_value", 0))),
    ]
    for label, value in rows:
        ws.append([label, value])
    for col, width in zip("AB", (34, 20)):
        ws.column_dimensions[col].width = width

    # Bar chart: received vs sold vs expenses vs purchases value
    chart_data_ws = wb.create_sheet("ChartData")
    chart_data_ws.append(["Category", "Value (KES)"])
    chart_rows = [
        ("Goods Received", float(summary.get("goods_received_value", 0))),
        ("Manual Stock Added", float(summary.get("manual_stock_added_value", 0))),
        ("Sales", float(summary.get("sales_value", 0))),
        ("Purchases", float(summary.get("total_purchases_value", 0))),
        ("Expenses", float(summary.get("total_expenses", 0))),
    ]
    for label, value in chart_rows:
        chart_data_ws.append([label, value])

    bar = BarChart()
    bar.title = "Value by Category (KES)"
    bar.y_axis.title = "KES"
    bar.x_axis.title = "Category"
    data_ref = Reference(chart_data_ws, min_col=2, min_row=1,
                          max_row=1 + len(chart_rows))
    cats_ref = Reference(chart_data_ws, min_col=1, min_row=2,
                          max_row=1 + len(chart_rows))
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.width, bar.height = 18, 10
    ws.add_chart(bar, "D2")

    # ── Top moving parts sheet (with its own bar chart) ─────────────────
    ws2 = wb.create_sheet("Top Moving Parts")
    ws2.append(["SKU", "Part Number", "Product", "Qty Moved", "Value Moved (KES)"])
    _style_header(ws2, ncols=5)
    for row in top_parts:
        ws2.append([
            row.get("sku", ""), row.get("part_number", "") or "",
            row.get("product_name", ""),
            row.get("quantity_moved", 0), float(row.get("value_moved", 0)),
        ])
    for col, width in zip("ABCDE", (16, 20, 36, 12, 18)):
        ws2.column_dimensions[col].width = width

    if top_parts:
        n = len(top_parts)
        bar2 = BarChart()
        bar2.title = "Top Moving Parts by Quantity"
        bar2.y_axis.title = "Qty Moved"
        data_ref2 = Reference(ws2, min_col=4, min_row=1, max_row=1 + n)
        cats_ref2 = Reference(ws2, min_col=3, min_row=2, max_row=1 + n)
        bar2.add_data(data_ref2, titles_from_data=True)
        bar2.set_categories(cats_ref2)
        bar2.width, bar2.height = 18, 10
        ws2.add_chart(bar2, "G2")

    # ── Raw stock movement ledger sheet ──────────────────────────────────
    ws3 = wb.create_sheet("Stock Movements")
    headers = ["Date", "Product ID", "Branch ID", "Qty Delta",
               "Qty After", "Reason", "Reference", "User ID", "Note"]
    ws3.append(headers)
    _style_header(ws3, ncols=len(headers))
    for m in stock_movements:
        ws3.append([
            str(m.get("created_at", "")), m.get("product_id", ""),
            m.get("branch_id", ""), m.get("quantity_delta", 0),
            m.get("quantity_after", 0), m.get("reason", ""),
            m.get("reference", ""), m.get("user_id", ""), m.get("note", ""),
        ])
    for i, header in enumerate(headers, start=1):
        ws3.column_dimensions[get_column_letter(i)].width = max(14, len(header) + 4)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_goods_received_excel(rows: list) -> bytes:
    """rows: list of app.analytics.schemas.GoodsReceivedRow — new stock
    added in the period, one row per product."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Goods Received"
    headers = ["Part No.", "Product", "SKU", "Qty Added", "Value Added (KES)", "Last Received"]
    ws.append(headers)
    _style_header(ws, ncols=len(headers))
    for r in rows:
        ws.append([
            r.part_number or "", r.product_name, r.sku,
            r.quantity_received, float(r.value_received),
            r.last_received_at.strftime("%Y-%m-%d %H:%M") if r.last_received_at else "",
        ])
    for col, width in zip("ABCDEF", (18, 34, 16, 12, 18, 18)):
        ws.column_dimensions[col].width = width

    if rows:
        n = len(rows)
        bar = BarChart()
        bar.title = "Goods Received by Part (Qty)"
        bar.y_axis.title = "Qty Added"
        data_ref = Reference(ws, min_col=4, min_row=1, max_row=1 + n)
        cats_ref = Reference(ws, min_col=2, min_row=2, max_row=1 + n)
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        bar.width, bar.height = 18, 10
        ws.add_chart(bar, "H2")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_stock_status_excel(report, customer_rows: list | None = None) -> bytes:
    """report: an app.analytics.schemas.StockStatusReport (or None if this
    call is only exporting customer_rows). customer_rows: a list of
    CustomerPurchaseRow, added as its own sheet when provided."""
    wb = Workbook()
    first_sheet_used = False

    if report is not None:
        ws = wb.active
        ws.title = "Out of Stock"
        headers = ["Category", "Part Name", "SKU", "Part Number", "Qty on Hand", "Reorder Point"]
        show_price = any(p.price_kes is not None for cat in report.categories for p in cat.out_of_stock)
        if show_price:
            headers.append("Price (KES)")
        ws.append(headers)
        _style_header(ws, ncols=len(headers))
        for cat in report.categories:
            for p in cat.out_of_stock:
                row = [cat.category_name, p.name, p.sku, p.part_number or "", p.quantity_on_hand, p.reorder_point]
                if show_price:
                    row.append(round(p.price_kes / 100, 2) if p.price_kes is not None else "")
                ws.append(row)
        for col, width in zip("ABCDEFG", (20, 34, 14, 16, 12, 14, 14)):
            ws.column_dimensions[col].width = width

        ws2 = wb.create_sheet("Low Stock")
        ws2.append(headers)
        _style_header(ws2, ncols=len(headers))
        for cat in report.categories:
            for p in cat.low_stock:
                row = [cat.category_name, p.name, p.sku, p.part_number or "", p.quantity_on_hand, p.reorder_point]
                if show_price:
                    row.append(round(p.price_kes / 100, 2) if p.price_kes is not None else "")
                ws2.append(row)
        for col, width in zip("ABCDEFG", (20, 34, 14, 16, 12, 14, 14)):
            ws2.column_dimensions[col].width = width
        first_sheet_used = True

    if customer_rows is not None:
        ws3 = wb.active if not first_sheet_used else wb.create_sheet("Customer Purchases")
        ws3.title = "Customer Purchases"
        headers3 = ["Customer", "Part Number", "Part / Description", "Total Qty",
                    "Total Value (KES)", "Times Purchased"]
        ws3.append(headers3)
        _style_header(ws3, ncols=len(headers3))
        for r in customer_rows:
            ws3.append([
                r.customer_name, getattr(r, "part_number", None) or "",
                r.description, float(r.total_quantity),
                round(r.total_value_kes / 100, 2), r.purchase_count,
            ])
        for col, width in zip("ABCDEF", (28, 20, 40, 12, 18, 16)):
            ws3.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
