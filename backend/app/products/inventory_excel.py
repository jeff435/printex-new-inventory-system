import io
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

STATUS_FILL = {
    "OUT_OF_STOCK": PatternFill(start_color="FDE2E2", end_color="FDE2E2", fill_type="solid"),
    "LOW_STOCK": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "IN_STOCK": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
}

STATUS_LABEL = {
    "OUT_OF_STOCK": "Out of stock",
    "LOW_STOCK": "Low stock",
    "IN_STOCK": "In stock",
}

# "Part No." is the manufacturer's catalogue number staff reorder against;
# "SKU" is Printex's own internal stock code. The export used to carry only
# the SKU, which made a printed inventory sheet useless for placing an order
# with a supplier.
HEADERS = [
    "SKU", "Part No.", "Product", "Branch", "On Hand", "Reserved",
    "Available", "Reorder Point", "Status",
]


def _write_sheet(ws, rows: list):
    ws.append(HEADERS)
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    for item in rows:
        status = item.get("stock_status", "")
        available = max(0, item.get("quantity_on_hand", 0) - item.get("quantity_reserved", 0))
        ws.append([
            item.get("sku", ""),
            item.get("part_number", "") or "—",
            item.get("product_name", ""),
            item.get("branch_name", ""),
            item.get("quantity_on_hand", 0),
            item.get("quantity_reserved", 0),
            available,
            item.get("reorder_point", 0),
            STATUS_LABEL.get(status, status),
        ])
        fill = STATUS_FILL.get(status)
        if fill:
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill

    widths = (16, 20, 34, 20, 10, 10, 10, 14, 14)
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def render_inventory_excel(all_items: list, low_stock_items: list, out_of_stock_items: list) -> bytes:
    """Builds a workbook for secretaries/directors to analyse stock levels.

    all_items: every inventory row matching the current filters.
    low_stock_items / out_of_stock_items: pre-split subsets for their own
    sheets so a secretary can open straight to what needs attention.
    """
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Printex Inventory Report"])
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.append([f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"])
    ws_summary.append([])
    ws_summary.append(["Metric", "Count"])
    _style_row = ws_summary.max_row
    for col in (1, 2):
        cell = ws_summary.cell(row=_style_row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws_summary.append(["Total items in view", len(all_items)])
    ws_summary.append(["Low stock", len(low_stock_items)])
    ws_summary.append(["Out of stock", len(out_of_stock_items)])
    ws_summary.column_dimensions["A"].width = 28
    ws_summary.column_dimensions["B"].width = 16

    ws_all = wb.create_sheet("All Inventory")
    _write_sheet(ws_all, all_items)

    ws_low = wb.create_sheet("Low Stock")
    _write_sheet(ws_low, low_stock_items)

    ws_out = wb.create_sheet("Out of Stock")
    _write_sheet(ws_out, out_of_stock_items)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
