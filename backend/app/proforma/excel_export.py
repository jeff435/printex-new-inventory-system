"""Excel (.xlsx) export for a single proforma invoice, via openpyxl."""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = "14151A"
GREEN = "2F8F4E"
LIGHT_GREY = "F5F6F8"

thin = Side(style="thin", color="E6E8EB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _kes(cents: int) -> float:
    return round(cents / 100, 2)


def render_proforma_excel(inv) -> bytes:
    """inv: an app.proforma.models.ProformaInvoice with .items loaded."""
    wb = Workbook()
    ws = wb.active
    ws.title = inv.pi_number[:31]

    ws.merge_cells("A1:E1")
    ws["A1"] = "PRINTEX ENGINEERS — PROFORMA INVOICE"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)

    ws["A3"] = "PI Number:"
    ws["B3"] = inv.pi_number
    ws["A4"] = "Customer:"
    ws["B4"] = inv.customer_name
    ws["A5"] = "Phone:"
    ws["B5"] = inv.customer_phone or "—"
    ws["A6"] = "Email:"
    ws["B6"] = inv.customer_email or "—"
    ws["A7"] = "Address:"
    ws["B7"] = getattr(inv, "customer_address", None) or "—"
    ws["A8"] = "Status:"
    ws["B8"] = (inv.status.value if hasattr(inv.status, "value") else inv.status).title()
    ws["A9"] = "Valid until:"
    ws["B9"] = inv.valid_until or "—"
    for row in range(3, 10):
        ws[f"A{row}"].font = Font(bold=True, color=NAVY)

    header_row = 11
    headers = ["#", "Part No.", "Description", "Qty", "Unit Price (KES)", "Line Total (KES)"]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    r = header_row + 1
    for idx, it in enumerate(inv.items, start=1):
        qty = float(it.quantity)
        values = [idx, getattr(it, "part_number", None) or "—", it.description, qty,
                  _kes(it.unit_price_kes), _kes(it.line_total_kes)]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = BORDER
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)
            if col >= 4:
                cell.alignment = Alignment(horizontal="right")
            else:
                # Long descriptions wrap inside the cell instead of spilling
                # across the columns beside them when the sheet is printed.
                cell.alignment = Alignment(vertical="top", wrap_text=(col == 3))
        r += 1

    r += 1
    totals = [("Subtotal", _kes(inv.subtotal_kes))]
    if inv.discount_kes:
        totals.append((f"Discount ({float(inv.discount_pct):g}%)", -_kes(inv.discount_kes)))
    totals.append(("VAT (16%)", _kes(inv.tax_kes)))
    totals.append(("Total", _kes(inv.total_kes)))

    for label, val in totals:
        ws.cell(row=r, column=5, value=label).font = Font(bold=(label == "Total"), color=NAVY)
        amount_cell = ws.cell(row=r, column=6, value=val)
        amount_cell.alignment = Alignment(horizontal="right")
        if label == "Total":
            amount_cell.font = Font(bold=True, size=12, color=GREEN)
            ws.cell(row=r, column=5).font = Font(bold=True, size=12, color=NAVY)
        r += 1

    if inv.notes:
        r += 1
        ws.cell(row=r, column=1, value="Notes:").font = Font(bold=True, color=NAVY)
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1, value=inv.notes)

    widths = [6, 20, 46, 8, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
