"""Excel (.xlsx) export for a single Purchase Order, via openpyxl — same
pattern as app.proforma.excel_export."""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

NAVY = "14151a"


def render_purchase_order_excel(purchase) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Order"

    ws["A1"] = "PRINTEX ENGINEERS — PURCHASE ORDER"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)

    status = purchase.status.value if hasattr(purchase.status, "value") else purchase.status
    rows = [
        ("PO Number:", purchase.purchase_number),
        ("Date:", purchase.created_at.strftime("%Y-%m-%d") if getattr(purchase, "created_at", None) else "—"),
        ("Supplier:", getattr(purchase, "supplier_name", None) or "—"),
        ("Status:", str(status).title()),
    ]
    r = 3
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, color=NAVY)
        ws.cell(row=r, column=2, value=value)
        r += 1

    header_row = r + 1
    headers = ["#", "Part No.", "Description", "Qty", "Unit Cost (KES)", "Line Total (KES)"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)

    r = header_row + 1
    total = 0
    for idx, item in enumerate(purchase.items, start=1):
        line_total = float(item.subtotal if item.subtotal is not None else (item.quantity * item.unit_cost))
        total += line_total
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=getattr(item, "product_part_number", None) or "—")
        ws.cell(row=r, column=3, value=getattr(item, "product_name", None) or item.product_id)
        ws.cell(row=r, column=4, value=item.quantity)
        ws.cell(row=r, column=5, value=float(item.unit_cost))
        ws.cell(row=r, column=6, value=line_total)
        r += 1

    ws.cell(row=r, column=5, value="TOTAL").font = Font(bold=True, color=NAVY)
    ws.cell(row=r, column=6, value=total).font = Font(bold=True, color=NAVY)

    if purchase.notes:
        r += 2
        ws.cell(row=r, column=1, value="Notes:").font = Font(bold=True, color=NAVY)
        ws.cell(row=r, column=2, value=purchase.notes)

    widths = [4, 18, 42, 8, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_supplier_parts_excel(supplier_name: str, parts) -> bytes:
    """parts: list of app.purchases.schemas.SupplierTaggedPart."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Supplier Parts"

    ws["A1"] = f"PARTS SUPPLIED BY {supplier_name.upper()}"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)

    headers = ["Part No.", "Name", "SKU", "Price (USD)"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)

    r = 4
    for p in parts:
        ws.cell(row=r, column=1, value=p.part_number or "—")
        ws.cell(row=r, column=2, value=p.name)
        ws.cell(row=r, column=3, value=p.sku)
        ws.cell(row=r, column=4, value=(p.price_usd / 100) if p.price_usd is not None else None)
        r += 1

    widths = [18, 42, 16, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
