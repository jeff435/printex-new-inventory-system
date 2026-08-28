"""PDF/Excel export for a supplier's tagged-parts list — everything ever
tagged as bought from a given supplier, independent of any single Purchase
Order. Mirrors the style of pdf.py / excel_export.py but for a plain list
of parts rather than an order with quantities and costs."""
import io
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

NAVY = colors.HexColor("#14151a")
LIGHT_GREY = colors.HexColor("#f5f6f8")
BORDER_GREY = colors.HexColor("#e6e8eb")

COMPANY_NAME = "PRINTEX ENGINEERS LIMITED"
COMPANY_ADDRESS_LINES = ["P.O BOX 5800-00200", "NAIROBI-KENYA"]


def _esc(value) -> str:
    if value is None:
        return ""
    return str(value).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def render_supplier_parts_pdf(supplier, parts) -> bytes:
    """supplier: app.purchases.models.Supplier
    parts: list of SupplierTaggedPart-shaped objects (product_id, name, sku,
    part_number, price_usd)."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm, topMargin=16 * mm, bottomMargin=16 * mm,
        title=f"Parts tagged with {supplier.name}",
    )
    styles = getSampleStyleSheet()
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=9, leading=12)
    hdr = ParagraphStyle("hdr", parent=small, textColor=colors.white, fontName="Helvetica-Bold")
    num_hdr = ParagraphStyle("num_hdr", parent=hdr)
    num = ParagraphStyle("num", parent=small)

    story = [
        Paragraph(f"<b>{_esc(COMPANY_NAME)}</b>", styles["Title"]),
    ]
    for line in COMPANY_ADDRESS_LINES:
        story.append(Paragraph(_esc(line), small))
    story.append(Spacer(1, 10))
    story.append(HRFlowable(width="100%", thickness=0.75, color=BORDER_GREY))
    story.append(Spacer(1, 8))

    story.append(Paragraph("<b>PARTS TAGGED TO SUPPLIER</b>", styles["Heading2"]))
    story.append(Paragraph(f"Supplier: <b>{_esc(supplier.name)}</b>", small))
    if getattr(supplier, "contact_person", None):
        story.append(Paragraph(f"Contact: {_esc(supplier.contact_person)}", small))
    story.append(Spacer(1, 12))

    data = [[
        Paragraph("Part No.", hdr), Paragraph("Name", hdr),
        Paragraph("SKU", hdr), Paragraph("Buying Price (USD)", num_hdr),
    ]]
    for p in parts:
        price = f"${(p.price_usd / 100):,.2f}" if getattr(p, "price_usd", None) is not None else "—"
        data.append([
            Paragraph(_esc(p.part_number) or "—", small),
            Paragraph(_esc(p.name), small),
            Paragraph(_esc(p.sku) or "—", small),
            Paragraph(price, num),
        ])

    tbl = Table(data, colWidths=[30 * mm, 70 * mm, 35 * mm, 35 * mm], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GREY]),
        ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GREY),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 6))
    story.append(Paragraph(f"Total parts: {len(parts)}", small))

    doc.build(story)
    return buf.getvalue()


def render_supplier_history_pdf(supplier, rows) -> bytes:
    """supplier: app.purchases.models.Supplier
    rows: list of SupplierPurchaseHistoryRow-shaped objects — what's actually
    been bought from this supplier (RECEIVED purchase orders only), as
    opposed to render_supplier_parts_pdf above (which is just 'could sell
    us')."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm, topMargin=16 * mm, bottomMargin=16 * mm,
        title=f"Purchase history — {supplier.name}",
    )
    styles = getSampleStyleSheet()
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=9, leading=12)
    hdr = ParagraphStyle("hdr", parent=small, textColor=colors.white, fontName="Helvetica-Bold")
    num_hdr = ParagraphStyle("num_hdr", parent=hdr)
    num = ParagraphStyle("num", parent=small)

    story = [
        Paragraph(f"<b>{_esc(COMPANY_NAME)}</b>", styles["Title"]),
    ]
    for line in COMPANY_ADDRESS_LINES:
        story.append(Paragraph(_esc(line), small))
    story.append(Spacer(1, 10))
    story.append(HRFlowable(width="100%", thickness=0.75, color=BORDER_GREY))
    story.append(Spacer(1, 8))

    story.append(Paragraph("<b>PARTS ACTUALLY BOUGHT FROM SUPPLIER</b>", styles["Heading2"]))
    story.append(Paragraph(f"Supplier: <b>{_esc(supplier.name)}</b>", small))
    if getattr(supplier, "contact_person", None):
        story.append(Paragraph(f"Contact: {_esc(supplier.contact_person)}", small))
    story.append(Paragraph("Received purchase orders only — draft/cancelled orders excluded.", small))
    story.append(Spacer(1, 12))

    data = [[
        Paragraph("Part No.", hdr), Paragraph("Name", hdr), Paragraph("SKU", hdr),
        Paragraph("Qty Bought", num_hdr), Paragraph("Total Spent (KES)", num_hdr),
        Paragraph("Last Bought", num_hdr),
    ]]
    total_spent = 0
    for r in rows:
        total_spent += r.total_spent_kes
        last = r.last_purchased_at[:10] if r.last_purchased_at else "—"
        data.append([
            Paragraph(_esc(r.part_number) or "—", small),
            Paragraph(_esc(r.name), small),
            Paragraph(_esc(r.sku) or "—", small),
            Paragraph(str(r.total_quantity), num),
            Paragraph(f"{(r.total_spent_kes / 100):,.2f}", num),
            Paragraph(last, num),
        ])

    tbl = Table(data, colWidths=[25 * mm, 50 * mm, 25 * mm, 22 * mm, 30 * mm, 22 * mm], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GREY]),
        ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GREY),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 8))
    story.append(Paragraph(f"Total parts: {len(rows)}", small))
    story.append(Paragraph(f"<b>Total spent: KES {(total_spent / 100):,.2f}</b>", small))

    doc.build(story)
    return buf.getvalue()


def render_supplier_history_excel(supplier, rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase History"

    ws["A1"] = "PRINTEX ENGINEERS — SUPPLIER PURCHASE HISTORY"
    ws["A1"].font = Font(bold=True, size=14, color="14151a")
    ws["A3"] = "Supplier:"
    ws["A3"].font = Font(bold=True, color="14151a")
    ws["B3"] = supplier.name
    ws["A4"] = "Received purchase orders only — draft/cancelled orders excluded."

    header_row = 6
    headers = ["#", "Part No.", "Name", "SKU", "Qty Bought", "Total Spent (KES)", "Last Bought"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="14151a")

    r = header_row + 1
    total_spent = 0
    for idx, row in enumerate(rows, start=1):
        total_spent += row.total_spent_kes
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=row.part_number or "—")
        ws.cell(row=r, column=3, value=row.name)
        ws.cell(row=r, column=4, value=row.sku or "—")
        ws.cell(row=r, column=5, value=row.total_quantity)
        ws.cell(row=r, column=6, value=row.total_spent_kes / 100)
        ws.cell(row=r, column=7, value=row.last_purchased_at[:10] if row.last_purchased_at else "—")
        r += 1

    r += 1
    ws.cell(row=r, column=5, value="Total:").font = Font(bold=True)
    ws.cell(row=r, column=6, value=total_spent / 100).font = Font(bold=True)

    widths = [4, 18, 42, 18, 14, 20, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_supplier_parts_excel(supplier, parts) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tagged Parts"

    ws["A1"] = "PRINTEX ENGINEERS — SUPPLIER TAGGED PARTS"
    ws["A1"].font = Font(bold=True, size=14, color="14151a")
    ws["A3"] = "Supplier:"
    ws["A3"].font = Font(bold=True, color="14151a")
    ws["B3"] = supplier.name

    header_row = 5
    headers = ["#", "Part No.", "Name", "SKU", "Buying Price (USD)"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="14151a")

    r = header_row + 1
    for idx, p in enumerate(parts, start=1):
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=p.part_number or "—")
        ws.cell(row=r, column=3, value=p.name)
        ws.cell(row=r, column=4, value=p.sku or "—")
        ws.cell(row=r, column=5, value=(p.price_usd / 100) if p.price_usd is not None else None)
        r += 1

    widths = [4, 18, 42, 18, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
